"""Плательщики, платежи, деректораты, настройки взносов."""
import io
from datetime import date
from decimal import Decimal
from enum import Enum
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from backend.application.schemas import (
    BudgetSettings, DataEntryContext, FacultyCreate, FacultyResponse, FacultyUpdate,
    GroupHint,
    PaginatedPayers, PayerCreate, PayerResponse, PayerUpdate,
    PayerWithDetailsResponse, PaymentCreate, PaymentResponse,
    PaymentSettingsCreate, PaymentSettingsResponse, PaymentSettingsUpdate,
    PaymentUpdate,
)
from backend.core.database import get_db
from backend.domain.academic import (
    LEVEL_LABELS, academic_year_label, academic_year_start, parse_level,
)
from backend.domain.models import (
    Faculty, Payer, Payment, PaymentSettings, PaymentStatus, SystemUser,
)
from backend.infrastructure.repositories import (
    AppSettingsRepository, AuditRepository, FacultyRepository, PayerRepository,
    PaymentRepository, PaymentSettingsRepository,
)
from backend.presentation.dependencies import (
    client_ip, require_any_role, require_operator,
)

router = APIRouter(tags=["Плательщики"])


class ArchiveFilter(str, Enum):
    """
    Что показывать из архива.

    Строка, а не Optional[bool]: тройное состояние булевым query-параметром
    по HTTP не выразить — «не передан» и «передан false» неразличимы,
    а axios вообще выбрасывает undefined из параметров запроса.
    """
    ACTIVE = "active"
    ARCHIVED = "archived"
    ALL = "all"


_ARCHIVE_FLAG = {
    ArchiveFilter.ACTIVE: False,
    ArchiveFilter.ARCHIVED: True,
    ArchiveFilter.ALL: None,
}

# Поля плательщика, которые уходят в ответ как есть.
_PLAIN_FIELDS = (
    "id", "last_name", "first_name", "middle_name", "date_of_birth",
    "email", "phone", "telegram", "vk",
    "is_budget", "stipend_amount", "budget_percent",
    "faculty_id", "group_name", "department", "admission_year",
    "status", "membership_start", "membership_end", "is_active", "notes",
    "created_at", "updated_at", "decryption_failed",
)


def serialize_payer(payer: Payer) -> dict:
    """
    Единственное место, где собирается ответ по плательщику.

    Раньше этот словарь был скопирован в четыре ручки, и новое поле
    приходилось дописывать в каждую копию.
    """
    data = {field: getattr(payer, field) for field in _PLAIN_FIELDS}
    data.update({
        "full_name": payer.full_name,
        "total_paid": payer.total_paid,
        "course": payer.computed_course,     # считается из года поступления
        "group_code": payer.group_code,      # код группы с актуальным курсом
        "education_level": payer.education_level or "bachelor",
        "is_archived": payer.is_archived,
    })
    return data


def _apply_updates(payer: Payer, data: dict) -> None:
    """Наложить изменения, приведя перечисления к строкам для String-колонок."""
    for field, value in data.items():
        if isinstance(value, Enum):
            value = value.value
        setattr(payer, field, value)


# ============== Деректораты ==============

@router.get("/faculties", response_model=list[FacultyResponse])
async def list_faculties(
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_any_role),
):
    return FacultyRepository(db).get_all(active_only=active_only)


@router.post("/faculties", response_model=FacultyResponse, status_code=status.HTTP_201_CREATED)
async def create_faculty(
    data: FacultyCreate,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_operator),
):
    return FacultyRepository(db).create(Faculty(name=data.name, short_name=data.short_name))


@router.put("/faculties/{faculty_id}", response_model=FacultyResponse)
async def update_faculty(
    faculty_id: int,
    data: FacultyUpdate,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_operator),
):
    repo = FacultyRepository(db)
    faculty = repo.get_by_id(faculty_id)
    if not faculty:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Деректорат не найден")

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(faculty, field, value)
    return repo.save(faculty)


@router.delete("/faculties/{faculty_id}")
async def delete_faculty(
    faculty_id: int,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_operator),
):
    """
    Удалить деректорат. Только пустой — иначе плательщики потеряли бы привязку.
    Если к нему кто-то привязан, предлагаем деактивацию.
    """
    repo = FacultyRepository(db)
    if not repo.get_by_id(faculty_id):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Деректорат не найден")

    linked = repo.count_payers(faculty_id)
    if linked:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"К деректорату привязано плательщиков: {linked}. "
            f"Его можно деактивировать — он исчезнет из списков, а данные останутся.",
        )

    repo.delete(faculty_id)
    return {"message": "Деректорат удалён"}


# ============== Суммы взносов ==============

@router.get("/payment-settings", response_model=list[PaymentSettingsResponse])
async def list_payment_settings(
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_any_role),
):
    return PaymentSettingsRepository(db).get_all()


@router.get("/payment-settings/current", response_model=PaymentSettingsResponse)
async def current_payment_settings(
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_any_role),
):
    item = PaymentSettingsRepository(db).get_current()
    if not item:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Суммы взносов ещё не заданы")
    return item


@router.post("/payment-settings", response_model=PaymentSettingsResponse,
             status_code=status.HTTP_201_CREATED)
async def create_payment_settings(
    data: PaymentSettingsCreate,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_operator),
):
    repo = PaymentSettingsRepository(db)
    if repo.get_by_year(data.academic_year):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Суммы на {data.academic_year} уже заданы",
        )
    return repo.create(PaymentSettings(**data.model_dump()))


@router.put("/payment-settings/{settings_id}", response_model=PaymentSettingsResponse)
async def update_payment_settings(
    settings_id: int,
    data: PaymentSettingsUpdate,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_operator),
):
    repo = PaymentSettingsRepository(db)
    item = repo.get_by_id(settings_id)
    if not item:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Настройки не найдены")

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(item, field, value)
    return repo.save(item)


@router.delete("/payment-settings/{settings_id}")
async def delete_payment_settings(
    settings_id: int,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_operator),
):
    if not PaymentSettingsRepository(db).delete(settings_id):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Настройки не найдены")
    return {"message": "Настройки удалены"}


# ============== Настройки бюджетников ==============

@router.get("/budget-settings", response_model=BudgetSettings)
async def get_budget_settings(
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_any_role),
):
    repo = AppSettingsRepository(db)
    return BudgetSettings(
        default_budget_percent=repo.get_value("default_budget_percent", "1") or "1",
        default_stipend_amount=repo.get_value("default_stipend_amount", "") or "",
    )


@router.put("/budget-settings", response_model=BudgetSettings)
async def update_budget_settings(
    data: BudgetSettings,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_operator),
):
    repo = AppSettingsRepository(db)
    repo.set("default_budget_percent", data.default_budget_percent, "Процент от стипендии по умолчанию")
    repo.set("default_stipend_amount", data.default_stipend_amount, "Сумма стипендии по умолчанию")
    return data


# ============== Контекст ввода данных ==============

@router.get("/data-entry-context", response_model=DataEntryContext)
async def data_entry_context(
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_any_role),
):
    """
    Что подставится в новую запись из настроек системы.

    Форма добавления плательщика показывает это прямо на месте: какой сейчас
    учебный год, какие суммы взносов действуют и какие значения по умолчанию
    стоят для бюджетников. Раньше настройки жили сами по себе, и по форме
    ввода было не понять, что именно из них применяется.
    """
    year_settings = PaymentSettingsRepository(db).get_current()
    app_settings = AppSettingsRepository(db)

    return DataEntryContext(
        academic_year=academic_year_label(),
        academic_year_start=academic_year_start(),
        fall_amount=year_settings.fall_amount if year_settings else None,
        spring_amount=year_settings.spring_amount if year_settings else None,
        year_total=year_settings.total_year_amount if year_settings else None,
        currency=year_settings.currency if year_settings else "RUB",
        has_payment_settings=year_settings is not None,
        default_budget_percent=app_settings.get_value("default_budget_percent", "1") or "1",
        default_stipend_amount=app_settings.get_value("default_stipend_amount", "") or "",
        faculties_count=len(FacultyRepository(db).get_all(active_only=True)),
    )


@router.get("/group-hints", response_model=list[GroupHint])
async def group_hints(
    faculty_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_any_role),
):
    """Уже заведённые группы — чтобы кафедра и уровень подставлялись сами."""
    return PayerRepository(db).group_hints(faculty_id=faculty_id)


# ============== Плательщики ==============

@router.get("/payers", response_model=PaginatedPayers)
async def list_payers(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    faculty_id: Optional[int] = None,
    status_filter: Optional[PaymentStatus] = Query(None, alias="status"),
    search: Optional[str] = Query(None, max_length=100),
    archive: ArchiveFilter = Query(ArchiveFilter.ACTIVE),
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_any_role),
):
    """Страница списка плательщиков. Фильтры и сортировка считаются в SQL."""
    payers, total = PayerRepository(db).list(
        skip=(page - 1) * per_page,
        limit=per_page,
        faculty_id=faculty_id,
        status=status_filter,
        search=search,
        archived=_ARCHIVE_FLAG[archive],
    )
    return PaginatedPayers(
        items=[serialize_payer(p) for p in payers],
        total=total,
        page=page,
        per_page=per_page,
        pages=max(1, (total + per_page - 1) // per_page),
    )


@router.get("/debtors", response_model=PaginatedPayers)
async def list_debtors(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    faculty_id: Optional[int] = None,
    search: Optional[str] = Query(None, max_length=100),
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_any_role),
):
    """Должники. Выпускники сюда не попадают — с них уже нечего требовать."""
    payers, total = PayerRepository(db).list(
        skip=(page - 1) * per_page,
        limit=per_page,
        faculty_id=faculty_id,
        search=search,
        archived=False,
        debtors_only=True,
    )
    return PaginatedPayers(
        items=[serialize_payer(p) for p in payers],
        total=total,
        page=page,
        per_page=per_page,
        pages=max(1, (total + per_page - 1) // per_page),
    )


# Символы, с которых Excel и LibreOffice начинают считать содержимое ячейки
# формулой. Значения приходят из карточек плательщиков, то есть их набирает
# человек, а выгрузку потом открывает бухгалтер — примечание вида
# `=HYPERLINK(...)` или DDE-вызов сработали бы у него на машине.
_FORMULA_STARTERS = ("=", "+", "-", "@", "\t", "\r")


def _cell(value):
    """Обезвредить значение перед записью в лист Excel."""
    if not isinstance(value, str) or not value.startswith(_FORMULA_STARTERS):
        return value
    # Апостроф в начале — принятый способ заставить Excel считать содержимое
    # текстом. В самой ячейке он не показывается.
    return "'" + value


@router.get("/payers/export")
async def export_payers(
    faculty_id: Optional[int] = None,
    status_filter: Optional[PaymentStatus] = Query(None, alias="status"),
    search: Optional[str] = Query(None, max_length=100),
    archive: ArchiveFilter = Query(ArchiveFilter.ACTIVE),
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_any_role),
):
    """Выгрузка в Excel."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    payers, _ = PayerRepository(db).list(
        skip=0, limit=10_000,
        faculty_id=faculty_id, status=status_filter, search=search,
        archived=_ARCHIVE_FLAG[archive],
    )
    faculties = {f.id: (f.short_name or f.name) for f in FacultyRepository(db).get_all(active_only=False)}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Плательщики"

    headers = [
        "№", "ФИО", "Деректорат", "Группа", "Курс", "Уровень", "Год поступл.", "Кафедра",
        "Email", "Телефон", "Telegram", "Бюджетник", "Стипендия", "%",
        "Статус", "Оплачено (₽)", "Д. рождения", "Примечание",
    ]
    widths = [5, 34, 18, 12, 7, 14, 13, 14, 26, 17, 15, 11, 12, 6, 14, 15, 14, 30]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F9788", end_color="1F9788", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, (title, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=idx, value=title)
        cell.font, cell.fill, cell.alignment = header_font, header_fill, center
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 30

    labels = {"paid": "Оплачено", "unpaid": "Не оплачено", "partial": "Частично", "exempt": "Освобождён"}

    for row_idx, payer in enumerate(payers, start=2):
        raw_status = payer.status.value if hasattr(payer.status, "value") else payer.status
        ws.append([_cell(value) for value in (
            row_idx - 1,
            payer.full_name,
            faculties.get(payer.faculty_id, ""),
            payer.group_code or "",
            payer.computed_course or "",
            LEVEL_LABELS[parse_level(payer.education_level)],
            payer.admission_year or "",
            payer.department or "",
            payer.email or "",
            payer.phone or "",
            payer.telegram or "",
            "Да" if payer.is_budget else "Нет",
            float(payer.stipend_amount) if payer.stipend_amount else "",
            float(payer.budget_percent) if payer.budget_percent else "",
            "В архиве" if payer.is_archived else labels.get(raw_status, raw_status),
            float(payer.total_paid or 0),
            str(payer.date_of_birth) if payer.date_of_birth else "",
            payer.notes or "",
        )])
        if row_idx % 2 == 0:
            fill = PatternFill(start_color="F0FAFA", end_color="F0FAFA", fill_type="solid")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"profpay_{date.today():%Y-%m-%d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@router.get("/payers/{payer_id}", response_model=PayerWithDetailsResponse)
async def get_payer(
    payer_id: int,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_any_role),
):
    payer = PayerRepository(db).get_by_id(payer_id)
    if not payer:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Плательщик не найден")

    data = serialize_payer(payer)
    data["faculty"] = payer.faculty
    data["payments"] = sorted(payer.payments, key=lambda p: (p.payment_date, p.id), reverse=True)
    return data


@router.post("/payers", response_model=PayerResponse, status_code=status.HTTP_201_CREATED)
async def create_payer(
    data: PayerCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_operator),
):
    if data.faculty_id and not FacultyRepository(db).get_by_id(data.faculty_id):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Деректорат не найден")

    payer = Payer(created_by=current_user.id)
    _apply_updates(payer, data.model_dump())

    created = PayerRepository(db).create(payer)
    AuditRepository(db).record(
        "create", "payer", created.id, current_user.id,
        f"Добавлен плательщик {created.full_name}", client_ip(request),
    )
    return serialize_payer(created)


@router.put("/payers/{payer_id}", response_model=PayerResponse)
async def update_payer(
    payer_id: int,
    data: PayerUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_operator),
):
    """
    Изменить плательщика.

    Порядок «расшифровать всё → наложить изменения → зашифровать всё»
    здесь принципиален. Раньше изменения накладывались прямо на зашифрованную
    запись, а потом шифровалось всё подряд — непереданные поля получали второй
    слой и становились нечитаемыми.
    """
    repo = PayerRepository(db)
    payer = db.query(Payer).filter(Payer.id == payer_id).first()
    if not payer:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Плательщик не найден")

    from backend.infrastructure.repositories import decrypt_payer
    decrypt_payer(payer)
    if payer.decryption_failed:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            "Часть данных этой записи зашифрована другим ключом. Сохранение отменено, "
            "чтобы не потерять их — проверьте ENCRYPTION_KEY в .env",
        )

    updates = data.model_dump(exclude_unset=True)
    if updates.get("faculty_id") and not FacultyRepository(db).get_by_id(updates["faculty_id"]):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Деректорат не найден")

    _apply_updates(payer, updates)
    saved = repo.save(payer)

    AuditRepository(db).record(
        "update", "payer", payer_id, current_user.id,
        f"Изменён плательщик {saved.full_name}: {', '.join(sorted(updates))}", client_ip(request),
    )
    return serialize_payer(saved)


@router.delete("/payers/{payer_id}")
async def delete_payer(
    payer_id: int,
    request: Request,
    hard: bool = Query(False, description="true — удалить безвозвратно вместе с платежами"),
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_operator),
):
    """По умолчанию запись прячется, а не стирается — восстановить можно."""
    repo = PayerRepository(db)
    ok = repo.delete(payer_id) if hard else repo.deactivate(payer_id)
    if not ok:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Плательщик не найден")

    AuditRepository(db).record(
        "delete" if hard else "deactivate", "payer", payer_id, current_user.id,
        None, client_ip(request),
    )
    return {"message": "Плательщик удалён" if hard else "Плательщик скрыт из списков"}


# ============== Платежи ==============

def recalculate_payer_status(db: Session, payer_id: int) -> None:
    """
    Пересчитать статус оплаты по фактическим платежам.

    Отдельная функция, потому что раньше статус выставлялся только при создании
    платежа: после удаления единственного платежа человек оставался «Оплачено»
    с нулевой суммой.

    Порог сравнивается с суммой за учебный год из настроек: полная сумма — PAID,
    что-то меньшее — PARTIAL, ноль — UNPAID. Если суммы на год не заданы,
    любой платёж считается полной оплатой (прежнее поведение).
    Освобождённых (exempt) не трогаем — это ручное решение.
    """
    payer = db.query(Payer).filter(Payer.id == payer_id).first()
    if payer is None or payer.status == PaymentStatus.EXEMPT:
        return

    total = PayerRepository(db).total_paid(payer_id)
    year_total = None
    current = PaymentSettingsRepository(db).get_current()
    if current:
        year_total = current.total_year_amount

    if total <= 0:
        new_status = PaymentStatus.UNPAID
    elif year_total and total < year_total:
        new_status = PaymentStatus.PARTIAL
    else:
        new_status = PaymentStatus.PAID

    if payer.status != new_status:
        payer.status = new_status
        db.commit()


@router.get("/payers/{payer_id}/payments", response_model=list[PaymentResponse])
async def list_payer_payments(
    payer_id: int,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_any_role),
):
    return PaymentRepository(db).get_by_payer(payer_id)


@router.post("/payments", response_model=PaymentResponse, status_code=status.HTTP_201_CREATED)
async def create_payment(
    data: PaymentCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_operator),
):
    if not db.query(Payer.id).filter(Payer.id == data.payer_id).first():
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Плательщик не найден")

    payment = Payment(created_by=current_user.id)
    _apply_updates(payment, data.model_dump())

    created = PaymentRepository(db).create(payment)
    recalculate_payer_status(db, data.payer_id)

    AuditRepository(db).record(
        "create", "payment", created.id, current_user.id,
        f"Платёж {created.amount} ₽", client_ip(request),
    )
    return created


@router.put("/payments/{payment_id}", response_model=PaymentResponse)
async def update_payment(
    payment_id: int,
    data: PaymentUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_operator),
):
    payment = db.query(Payment).filter(Payment.id == payment_id).first()
    if not payment:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Платёж не найден")

    from backend.infrastructure.repositories import decrypt_payment
    decrypt_payment(payment)
    _apply_updates(payment, data.model_dump(exclude_unset=True))

    saved = PaymentRepository(db).save(payment)
    recalculate_payer_status(db, saved.payer_id)

    AuditRepository(db).record(
        "update", "payment", payment_id, current_user.id, None, client_ip(request)
    )
    return saved


@router.delete("/payments/{payment_id}")
async def delete_payment(
    payment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: SystemUser = Depends(require_operator),
):
    # payer_id нужен до удаления — после него связь уже не прочитать.
    payer_id = db.query(Payment.payer_id).filter(Payment.id == payment_id).scalar()
    if not PaymentRepository(db).delete(payment_id):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Платёж не найден")

    if payer_id is not None:
        recalculate_payer_status(db, payer_id)

    AuditRepository(db).record(
        "delete", "payment", payment_id, current_user.id, None, client_ip(request)
    )
    return {"message": "Платёж удалён"}
